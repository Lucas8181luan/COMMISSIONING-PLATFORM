"""
excel_generator.py
Gera o arquivo Excel (.xlsx) da Relação de Leads / Dashboard, a partir dos
mesmos dados usados para o PDF.

Colocar em: APP/backend/excel_generator.py  (arquivo NOVO)
"""

import io
from datetime import datetime
from gsheet_client import agora_brasilia

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRETO = "0B0B0B"
CIANO = "00E5FF"
BRANCO = "FFFFFF"


def gerar_excel_dashboard(headers: list, rows: list, titulo: str = "RELAÇÃO DE LEADS",
                           cor_header_hex: str = PRETO, cor_texto_header_hex: str = CIANO) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    borda_fina = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    cel_titulo = ws.cell(row=1, column=1, value=titulo)
    cel_titulo.font = Font(bold=True, size=14, color="000000")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    cel_sub = ws.cell(row=2, column=1, value=f"Gerado em {agora_brasilia().strftime('%d/%m/%Y %H:%M')}")
    cel_sub.font = Font(italic=True, size=9, color="666666")

    linha_cabecalho = 4
    for col_idx, texto in enumerate(headers, start=1):
        cel = ws.cell(row=linha_cabecalho, column=col_idx, value=texto)
        cel.font = Font(bold=True, color=cor_texto_header_hex, size=10)
        cel.fill = PatternFill("solid", fgColor=cor_header_hex)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = borda_fina

    status_col_idx = None
    for i, h in enumerate(headers, start=1):
        if h.strip().upper() == "STATUS":
            status_col_idx = i
            break

    for r_offset, row in enumerate(rows):
        linha = linha_cabecalho + 1 + r_offset
        linha_vazia = not any(str(c).strip() for c in row)
        eh_ultima = (r_offset == len(rows) - 1)

        for col_idx in range(1, len(headers) + 1):
            valor = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            cel = ws.cell(row=linha, column=col_idx, value=valor)
            cel.border = borda_fina
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if linha_vazia and eh_ultima:
                cel.fill = PatternFill("solid", fgColor=cor_header_hex)
                cel.font = Font(bold=True, color=cor_texto_header_hex)
            elif status_col_idx and col_idx == status_col_idx:
                if str(valor).strip().upper() == "SEM TURMA":
                    cel.fill = PatternFill("solid", fgColor="CC3939")
                else:
                    cel.fill = PatternFill("solid", fgColor="8CC78C")
                cel.font = Font(bold=True, color=BRANCO)

    # Largura de colunas: LOCAL/CURSO mais largas
    for col_idx, h in enumerate(headers, start=1):
        nome = h.strip().upper()
        letra = get_column_letter(col_idx)
        if nome in ("LOCAL", "CURSO", "CURSOS"):
            ws.column_dimensions[letra].width = 42
        elif nome == "DATA DE INÍCIO":
            ws.column_dimensions[letra].width = 24
        else:
            ws.column_dimensions[letra].width = 13

    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=3)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
